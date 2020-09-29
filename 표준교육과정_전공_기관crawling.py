from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from openpyxl import Workbook, load_workbook
import time

wb  = Workbook()
sheet1 = wb.active
sheet1.title = '표준교육과정 과목중심' #시트명
sheet1.cell(row=1, column=2).value = "해당기관"

driver = webdriver.Chrome()

kconti_url = 'https://www.cb.or.kr/creditbank/stdPro/nStdPro1_1.do'

driver.get(kconti_url)
driver.maximize_window()
#6-2 경영학
search_button = driver.find_element_by_css_selector("#contents > div.innerContView > div.stdProtResult > div > ul:nth-child(6) > li:nth-child(2) > a")
major_name = search_button.text
sheet1.cell(row=1, column=1).value = major_name
search_button.click()
time.sleep(1)
#두번째 버튼 - 해당기관
search_button_uni = driver.find_element_by_css_selector("#frm > div.innerContView > div.listDateWrap01 > ul > li:nth-child(1) > div.btnWrap > a:nth-child(2)")
search_button_uni.click()

row = 2
for page in range(1, 10):
    all_list = driver.find_element_by_css_selector('div.listDateWrap01.mt30')
    select_list = all_list.find_elements_by_css_selector('li')

    for item in select_list:
        ki_gwan = item.find_element_by_css_selector('a').text

        sheet1.cell(row=row, column=2).value = ki_gwan

        row = row + 1
    time.sleep(1)
    if page % 10 == 0:
        driver.find_element_by_link_text('다음 페이지로 이동').click()
    else:
        next = str(page + 1)
        try:
            driver.find_element_by_link_text(next).click()
        except NoSuchElementException: # 마지막 다음 page는 없으므로 예외상황 발생되므로 ...^^ 꼼수
            break

driver.quit()
wb.save("./excel_folder/" + major_name + "기관" + ".xlsx")
