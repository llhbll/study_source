from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook

wb  = Workbook()
sheet1 = wb.active
sheet1.title = '표준교육과정 과목중심' #시트명
sheet1.cell(row=1, column=2).value = "과목명"
sheet1.cell(row=1, column=3).value = '전공구분'
sheet1.cell(row=1, column=4).value = '강의시간'
sheet1.cell(row=1, column=5).value = '실습시간'

driver = webdriver.Chrome()
# driver2 = webdriver.Chrome()

kconti_url = 'https://www.cb.or.kr/creditbank/stdPro/nStdPro1_1.do'

def detail_work():

    all_list = driver.find_element_by_css_selector('div.listDateWrap01')
    select_list = all_list.find_elements_by_css_selector('li')

    for item in select_list:
        major_flag_s = item.find_element_by_css_selector('em').text
        subject = item.find_element_by_css_selector('a').text
        lecture_time = item.find_elements_by_css_selector('span')[2].text
        practice_time = item.find_elements_by_css_selector('span')[3].text

        sheet1.cell(row=row, column=1).value = major_name
        sheet1.cell(row=row, column=2).value = subject
        sheet1.cell(row=row, column=3).value = major_flag_s
        sheet1.cell(row=row, column=4).value = lecture_time
        sheet1.cell(row=row, column=5).value = practice_time

driver.get(kconti_url)

all_list = driver.find_element_by_css_selector('div.stdProtResult')
major_list = all_list.find_elements_by_css_selector('li')

row = 1
for major_item in major_list:

    major_name = major_item.find_element_by_css_selector('a').text

    # main_page = driver.current_window_handle

    search_button = major_item.find_element_by_css_selector('a')
    ret = search_button.send_keys(Keys.CONTROL + "\n")
    # driver.switch_to.window(driver.window_handles[1])

    detail_work()
    driver.back()
    row = row + 1
    # driver.close()
    # driver.switch_to.window(driver.window_handles[0])

wb.save(major_name + ".xlsx")

driver.quit()


