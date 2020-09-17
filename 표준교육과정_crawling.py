from selenium import webdriver

from openpyxl import Workbook, load_workbook

wb  = Workbook()
sheet1 = wb.active
sheet1.title = '표준교육과정 과목중심' #시트명
sheet1.cell(row=1, column=2).value = "과목명"
sheet1.cell(row=1, column=3).value = '전공구분'
sheet1.cell(row=1, column=4).value = '강의시간'
sheet1.cell(row=1, column=5).value = '실습시간'

driver = webdriver.Chrome()

kconti_url = 'https://www.cb.or.kr/creditbank/stdPro/nStdPro1_1.do'

driver.get(kconti_url)

search_button = driver.find_element_by_css_selector("#contents > div.innerContView > div.stdProtResult > div > ul:nth-child(6) > li:nth-child(2) > a")
major_name = search_button.text
search_button.click()


all_list = driver.find_element_by_css_selector('div.listDateWrap01')
select_list = all_list.find_elements_by_css_selector('li')

row = 2

for item in select_list:
    jungong_flag_s = item.find_element_by_css_selector('em').text
    subject = item.find_element_by_css_selector('a').text
    lecture_time = item.find_elements_by_css_selector('span')[2].text
    practice_time = item.find_elements_by_css_selector('span')[3].text

    sheet1.cell(row=row, column=2).value = subject
    sheet1.cell(row=row, column=3).value = jungong_flag_s
    sheet1.cell(row=row, column=4).value = lecture_time
    sheet1.cell(row=row, column=5).value = practice_time
    row = row + 1

driver.quit()
wb.save(major_name + ".xlsx")
# row = 2
#
# for page in range(1, int(page_cnt) + 1):
#     list = driver.find_elements_by_css_selector('#main-area > div:nth-child(7) > table > tbody > tr')
#     for item in list:
#         title = item.find_element_by_css_selector('a.article').text.strip()
#         # title = re.sub('[^0-9a-zA-Zㄱ-힗]', '', title)
#         if any(format in title for format in except_str):
#             continue
#         strip_title = str(title).replace(' ', '')
#         aa = title.find(keyword)
#         if strip_title.find(keyword.replace(' ', '')) == -1: #양쪽다 공백없이 키워드 비교
#             continue
#         writer = item.find_element_by_css_selector('a.m-tcol-c').text.strip()
#         ddate = item.find_element_by_css_selector('td.td_date').text.strip()
#         link = item.find_element_by_css_selector('a.article').get_attribute('href')
#
#         time.sleep(2)
#         driver2.get(link) #로긴하고 상세페이지 들어가면 모두 접근가능할 줄 알았는데 중고나라회원에 허용한 page는 접근이 안됨! 추후 방법을 찾아야 함. 현재 실력으로는 불가
#         time.sleep(2)     # 리턴data에 noindex, nofollow 그래서 안되는것으로 파악... 그래도 우회하는 방법이 없을려나???
#
#         try:
#            driver2.switch_to.frame("cafe_main")
#            cost = driver2.find_element_by_css_selector('span.cost').text
#         except NoSuchElementException:
#             cost = 'X'
#
#         sheet1.cell(row=row, column=2).value = title
#         sheet1.cell(row=row, column=2).hyperlink = link
#         sheet1.cell(row=row, column=3).value = writer
#         sheet1.cell(row=row, column=4).value = ddate
#         sheet1.cell(row=row, column=5).value = cost
#
#         row = row + 1
#
#
#     if page % 10 == 0:
#         driver.find_element_by_link_text('다음').click()
#     else:
#         next = str(page + 1)
#         driver.find_element_by_link_text(next).click()

# # driver2.quit()
# wb.save(keyword + ".xlsx")
#
#         #item.find_element_by_css_selector('a').click()  #가격을 구하기 위해서






