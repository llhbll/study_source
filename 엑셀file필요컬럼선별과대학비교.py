import os
from glob import glob
from openpyxl import load_workbook
from openpyxl import Workbook

path = r'E:\기본자료\평생교육원정보\타대학현황\수입지출'
file_list = glob(r'E:\기본자료\평생교육원정보\타대학현황\수입지출\*수입지출 현황.xlsx')
# glob('*.xls')

uni_dic = {'성신':[3, '성신여자대학교부설평생교육원'], '동덕':[4, '동덕여자대학교부설평생교육원'],
           '덕성':[5, '덕성여자대학교부설평생교육원'], '서울':[6, '서울여자대학교부설평생교육원'],
           '숙명':[7, '숙명여자대학교 미래교육원', '숙명여자대학교부설평생교육원'],
           '이화':[8, '이화여자대학교부설평생교육원', '이화여자대학교 글로벌미래평생교육원'] }
year_dic = {'2015':[5, '2015'], '2016':[6, '2016'], '2017':[7, '2017'], '2018':[8, '2018'], '2019':[9, '2019']} # row 행
# index

wb_new  = Workbook()
sheet1 = wb_new.active
sheet2 = wb_new.create_sheet('주요대학 수입지출 현황')

sheet1.title = '타대학 년도별 수입지출 현황' #시트명
sheet1.cell(row=1, column=1).value = "결산년도"
sheet1.cell(row=1, column=2).value = "기관유형"
sheet1.cell(row=1, column=3).value = '공시기관'
sheet1.cell(row=1, column=4).value = '수강료수입'
sheet1.cell(row=1, column=5).value = '국고보조금'
sheet1.cell(row=1, column=6).value = '기타수입'
sheet1.cell(row=1, column=7).value = '수입합계'
sheet1.cell(row=1, column=8).value = '인건비'
sheet1.cell(row=1, column=9).value = '관리운영비'
sheet1.cell(row=1, column=10).value = '연구학생경비'
sheet1.cell(row=1, column=11).value = '기타비용'
sheet1.cell(row=1, column=12).value = '지출합계'

sheet2.cell(row=4, column=3).value = uni_dic['성신'][1] # 명칭
sheet2.cell(row=4, column=4).value = uni_dic['동덕'][1]
sheet2.cell(row=4, column=5).value = uni_dic['덕성'][1]
sheet2.cell(row=4, column=6).value = uni_dic['서울'][1]
sheet2.cell(row=4, column=7).value = uni_dic['숙명'][1]
sheet2.cell(row=4, column=8).value = uni_dic['이화'][1]
sheet2.cell(row=5, column=2).value = year_dic['2015'][1]
sheet2.cell(row=6, column=2).value = year_dic['2016'][1]
sheet2.cell(row=7, column=2).value = year_dic['2017'][1]
sheet2.cell(row=8, column=2).value = year_dic['2018'][1]
sheet2.cell(row=9, column=2).value = year_dic['2019'][1]

results = []

def pick_uni_ins(var_year, var_uni, var_in_sum):
    for uni_key, uni_data_list in uni_dic.items(): # uni_data_list = [7, '숙명여자대학교 미래교육원', '숙명여자대학교부설평생교육원']
        for uni_data in uni_data_list: # uni_data 차례대로 7, '숙명여자대학교 미래교육원', '숙명여자대학교부설평생교육원'
            if var_uni == uni_data:
                for year_key, year_data in year_dic.items(): # year_data = [5, "2015"]
                    if var_year == year_data[1]: # year_data 차례대로  5, "2015"
                        sheet2.cell(row=year_data[0], column=uni_data_list[0]).value = var_in_sum


for file_name_raw in file_list:

    file_name = file_name_raw
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb['결산(8월)']

    for row in ws.rows:
        result = []
        if row[0].value == '공시년도':
            continue
        result.append(str(int(row[0].value) -1)) #공시년도 -> -1년 하여 결산년도로 변경
        result.append(row[2].value) #기관유형
        result.append(row[3].value) #공시기관
        result.append(row[5].value) #수강료수입
        result.append(row[9].value) #국고보조금 수입
        result.append(row[11].value) #기타수입
        in_sum = float(row[5].value) + float(row[9].value) + float(row[11].value)
        result.append(str(in_sum))
        result.append(row[14].value) #인건비
        result.append(row[17].value) #관리운영비
        result.append(row[20].value) #연구학생경비
        result.append(row[26].value)  #기타비용
        out_sum = float(row[14].value) + float(row[17].value) + float(row[20].value) + float(row[26].value)
        result.append(str(out_sum))

        results.append(result)

        pick_uni_ins(str(int(row[0].value) -1), row[3].value, str(in_sum))
#        print(row[3].value)
#print(results)


for i in results:
    sheet1.append(i)

wb_new.save(r"E:\기본자료\평생교육원정보\타대학현황\수입지출\results.xlsx")